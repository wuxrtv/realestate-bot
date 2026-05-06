"""
Excel parser for Toni bot admin panel.
Handles multi-sheet workbooks, auto-detects unit number columns, builds searchable indexes.
"""

import io
import re
from typing import Any

import openpyxl

_UNIT_KEYS = frozenset({
    "unit", "unit no", "unit #", "unit number", "unit_number", "unit id",
    "юнит", "юнит №", "номер юнита", "номер", "квартира", "квартира №",
    "apartment", "flat", "no", "№",
})

_UNIT_RE = re.compile(r"^\d{3,5}$")


def normalize_project_name(filename: str) -> str:
    """'Breez Tower Units v2.xlsx' → 'Breez Tower'"""
    name = re.sub(r"\.(xlsx?|xls|csv)$", "", filename, flags=re.IGNORECASE)
    name = re.sub(r"[_]+", " ", name.strip())
    name = re.sub(
        r"\s+(v\d+|final|new|update|updated|rev\d*|\(\d+\)|\d{4}[-_]\d+)$",
        "",
        name,
        flags=re.IGNORECASE,
    )
    return name.strip() or "Проект"


def parse_excel(file_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    """
    Parse all non-empty sheets from an Excel file.
    Returns {sheet_name: [row_dicts]} using first non-empty row as headers.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw = list(ws.iter_rows(values_only=True))

        non_empty = [
            [c if c is not None else "" for c in row]
            for row in raw
            if any(c is not None and c != "" for c in row)
        ]

        if len(non_empty) < 2:
            continue

        # Detect header row: first row where ≥50% non-empty cells are strings
        header_idx = 0
        for i, row in enumerate(non_empty[:5]):
            filled = [v for v in row if v != ""]
            str_cells = sum(1 for v in filled if isinstance(v, str))
            if filled and str_cells / len(filled) >= 0.5:
                header_idx = i
                break

        header_row = non_empty[header_idx]
        headers: list[str] = []
        seen: dict[str, int] = {}
        for i, val in enumerate(header_row):
            h = str(val).strip() if val != "" else f"col_{i}"
            if h in seen:
                seen[h] += 1
                h = f"{h}_{seen[h]}"
            else:
                seen[h] = 0
            headers.append(h)

        rows: list[dict[str, Any]] = []
        for row in non_empty[header_idx + 1 :]:
            row_dict: dict[str, Any] = {
                headers[i]: (row[i] if i < len(row) else "")
                for i in range(len(headers))
            }
            if any(v != "" and v is not None for v in row_dict.values()):
                rows.append(row_dict)

        if rows:
            result[sheet_name] = rows

    wb.close()
    return result


def _clean(v: Any) -> str:
    """Normalize a cell value to a clean string."""
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _detect_unit_col(rows: list[dict]) -> str | None:
    """Return the column name most likely containing unit numbers."""
    if not rows:
        return None
    headers = list(rows[0].keys())

    for h in headers:
        if h.lower().strip() in _UNIT_KEYS:
            return h

    sample = rows[: min(15, len(rows))]
    for h in headers:
        vals = [_clean(r.get(h)) for r in sample if _clean(r.get(h))]
        if not vals:
            continue
        hits = sum(1 for v in vals if _UNIT_RE.match(v))
        if hits / len(vals) >= 0.6:
            return h

    return None


def build_unit_index(sheets_data: dict[str, list[dict]]) -> dict[str, dict[str, Any]]:
    """Build {unit_number: {_sheet, col: val, ...}} from all sheets."""
    index: dict[str, dict[str, Any]] = {}

    for sheet_name, rows in sheets_data.items():
        unit_col = _detect_unit_col(rows)

        for row in rows:
            unit_num = None

            if unit_col and unit_col in row:
                unit_num = _clean(row[unit_col])
                if not _UNIT_RE.match(unit_num or ""):
                    unit_num = None

            if not unit_num:
                for val in row.values():
                    s = _clean(val)
                    if _UNIT_RE.match(s):
                        unit_num = s
                        break

            if unit_num:
                index[unit_num] = {
                    "_sheet": sheet_name,
                    **{k: _clean(v) for k, v in row.items()},
                }

    return index


def diff_unit_indexes(old: dict, new: dict) -> dict:
    """Return {added, removed, changed} between two unit indexes."""
    added = {k: new[k] for k in new if k not in old}
    removed = {k: old[k] for k in old if k not in new}
    changed: dict[str, dict] = {}

    for k in old:
        if k not in new:
            continue
        diffs: dict[str, tuple] = {}
        for field in set(old[k]) | set(new[k]):
            if field == "_sheet":
                continue
            if str(old[k].get(field, "")) != str(new[k].get(field, "")):
                diffs[field] = (old[k].get(field, ""), new[k].get(field, ""))
        if diffs:
            changed[k] = diffs

    return {"added": added, "removed": removed, "changed": changed}


def format_diff_report(diff: dict, project_name: str) -> str:
    added, removed, changed = diff["added"], diff["removed"], diff["changed"]

    if not added and not removed and not changed:
        return f"✅ {project_name} — изменений нет, данные идентичны."

    parts = [f"📊 Изменения в проекте *{project_name}*:\n"]

    if added:
        parts.append(f"➕ Новых юнитов: {len(added)}")
        parts.extend(f"  • {u}" for u in list(added.keys())[:5])
        if len(added) > 5:
            parts.append(f"  ...ещё {len(added) - 5}")

    if removed:
        parts.append(f"\n➖ Удалено: {len(removed)}")
        parts.extend(f"  • {u}" for u in list(removed.keys())[:5])
        if len(removed) > 5:
            parts.append(f"  ...ещё {len(removed) - 5}")

    if changed:
        parts.append(f"\n✏️ Изменено: {len(changed)}")
        for u, fields in list(changed.items())[:5]:
            parts.append(f"  • {u}: {', '.join(fields.keys())}")
        if len(changed) > 5:
            parts.append(f"  ...ещё {len(changed) - 5}")

    return "\n".join(parts)


def format_unit_card(unit_num: str, data: dict, project_name: str) -> str:
    """Format unit data as a readable Telegram message."""
    sheet = data.get("_sheet", "")
    lines = [f"🏢 *Юнит {unit_num}* — {project_name}"]
    if sheet:
        lines.append(f"📍 {sheet}")

    skip = {"_sheet"}
    for k, v in data.items():
        if k in skip or not v or v in ("None", "nan", ""):
            continue
        lines.append(f"• {k}: {v}")

    return "\n".join(lines)
